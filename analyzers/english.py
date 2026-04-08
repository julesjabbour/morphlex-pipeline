"""English morphological analyzer using spaCy, MorphoLex, and MorphyNet."""

import os
import re
from pathlib import Path

import spacy
from openpyxl import load_workbook

# Initialize spaCy with the large English model
try:
    _nlp = spacy.load('en_core_web_lg')
except OSError:
    _nlp = None

# MorphoLex and MorphyNet data paths
_MORPHOLEX_PATH = Path('/mnt/pgdata/morphlex/MorphoLex-en/')
_MORPHYNET_PATH = Path('/mnt/pgdata/morphlex/MorphyNet/eng/')

# Cache for MorphoLex data
_morpholex_cache = None
_morphynet_cache = None


def _load_morpholex_data() -> dict:
    """
    Load MorphoLex data from all xlsx files in the MorphoLex-en directory.

    Returns:
        Dict mapping word -> MorphoLexSegm value
    """
    global _morpholex_cache

    if _morpholex_cache is not None:
        return _morpholex_cache

    _morpholex_cache = {}

    if not _MORPHOLEX_PATH.exists():
        print(f"WARNING: MorphoLex directory not found at {_MORPHOLEX_PATH} - using spaCy fallback")
        return _morpholex_cache

    # Find all xlsx files
    for xlsx_file in _MORPHOLEX_PATH.glob('*.xlsx'):
        try:
            wb = load_workbook(xlsx_file, read_only=True, data_only=True)

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Find column indices
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                if header_row is None:
                    continue

                word_col = None
                segm_col = None

                for idx, col_name in enumerate(header_row):
                    if col_name and 'word' in str(col_name).lower():
                        word_col = idx
                    if col_name and 'morpholexsegm' in str(col_name).lower():
                        segm_col = idx

                if word_col is None or segm_col is None:
                    continue

                # Read data rows
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > max(word_col, segm_col):
                        word = row[word_col]
                        segm = row[segm_col]
                        if word and segm:
                            _morpholex_cache[str(word).lower()] = str(segm)

            wb.close()
        except Exception as e:
            print(f"MorphoLex file load error ({xlsx_file}): {e}")
            continue

    return _morpholex_cache


def _load_morphynet_data() -> dict:
    """
    Load MorphyNet derivation data from TSV files.

    Returns:
        Dict mapping word -> derivation info
    """
    global _morphynet_cache

    if _morphynet_cache is not None:
        return _morphynet_cache

    _morphynet_cache = {}

    if not _MORPHYNET_PATH.exists():
        return _morphynet_cache

    # Find all TSV files
    for tsv_file in _MORPHYNET_PATH.glob('*.tsv'):
        try:
            with open(tsv_file, 'r', encoding='utf-8') as f:
                for line in f:
                    parts = line.strip().split('\t')
                    if len(parts) >= 3:
                        # Typical format: source_word, target_word, derivation_type
                        target_word = parts[1] if len(parts) > 1 else ''
                        deriv_type = parts[2] if len(parts) > 2 else ''
                        if target_word and deriv_type:
                            _morphynet_cache[target_word.lower()] = deriv_type
        except Exception as e:
            print(f"MorphyNet file load error ({tsv_file}): {e}")
            continue

    return _morphynet_cache


def _parse_morpholex_segm(segm: str) -> dict:
    """
    Parse MorphoLex segmentation notation.

    Format examples: {<un`<(happy)>} where:
    - () = root
    - <> = affixes (< prefix, > suffix)
    - ` = morpheme boundary

    Args:
        segm: MorphoLexSegm string

    Returns:
        Dict with 'root', 'prefixes', 'suffixes', 'components'
    """
    result = {
        'root': None,
        'prefixes': [],
        'suffixes': [],
        'components': []
    }

    if not segm:
        return result

    # Remove outer braces if present
    segm = segm.strip('{}')

    # Find root (in parentheses)
    root_match = re.search(r'\(([^)]+)\)', segm)
    if root_match:
        result['root'] = root_match.group(1)

    # Split by morpheme boundaries and parse
    # Prefixes come before the root and are marked with < at the start
    # Suffixes come after the root and are marked with > at the end

    # Find all morphemes
    # Pattern: < marks start of prefix, > marks end of suffix

    # Extract prefixes (text before root, between < and `)
    prefix_pattern = r'<([^<>`()]+)'
    for match in re.finditer(prefix_pattern, segm):
        prefix = match.group(1).strip('`')
        if prefix and '(' not in prefix:
            result['prefixes'].append(prefix)

    # Extract suffixes (text after root, ending with >)
    suffix_pattern = r'\)([^<>()]+)>'
    suffix_match = re.search(suffix_pattern, segm)
    if suffix_match:
        suffix_text = suffix_match.group(1)
        # Split by ` for multiple suffixes
        for s in suffix_text.split('`'):
            s = s.strip()
            if s:
                result['suffixes'].append(s)

    # Build components list
    result['components'] = result['prefixes'] + ([result['root']] if result['root'] else []) + result['suffixes']

    return result


def _get_derivation_type(prefixes: list, suffixes: list, morphynet_type: str = None) -> str:
    """
    Determine derivation type based on affixes and MorphyNet data.

    Args:
        prefixes: List of prefixes
        suffixes: List of suffixes
        morphynet_type: Derivation type from MorphyNet if available

    Returns:
        Derivation type string
    """
    if morphynet_type:
        return morphynet_type

    has_prefix = len(prefixes) > 0
    has_suffix = len(suffixes) > 0

    if has_prefix and has_suffix:
        return 'prefix+suffix'
    elif has_prefix:
        return 'prefix'
    elif has_suffix:
        return 'suffix'

    return None


def _classify_morph_type(root: str, prefixes: list, suffixes: list, components: list) -> str:
    """
    Classify morphological type based on structure.

    Returns: ROOT, DERIVATION, COMPOUND, COMPOUND_DERIVATION, OTHER, UNKNOWN
    """
    has_root = bool(root)
    has_affixes = bool(prefixes) or bool(suffixes)
    # Compounds have multiple root-level components (more than just root + affixes)
    is_compound = len(components) > 3 if components else False

    if is_compound and has_affixes:
        return 'COMPOUND_DERIVATION'
    elif is_compound:
        return 'COMPOUND'
    elif has_affixes and has_root:
        return 'DERIVATION'
    elif has_root and not has_affixes:
        return 'ROOT'
    elif has_root:
        return 'OTHER'
    else:
        return 'UNKNOWN'


def _fix_pos_tag(word: str, pos: str, lemma: str) -> str:
    """
    Fix common spaCy POS tagging errors.

    Problem 5: spaCy incorrectly tags common nouns as PROPN.
    """
    # List of words commonly mistagged as PROPN
    common_nouns = {
        'dictionary', 'book', 'water', 'fire', 'hand', 'eye', 'stone',
        'heart', 'sun', 'moon', 'tree', 'blood', 'house', 'word', 'name',
        'day', 'night', 'year', 'time', 'man', 'woman', 'child', 'world'
    }

    word_lower = word.lower()

    # If tagged as PROPN but is a common word, fix to NOUN
    if pos == 'PROPN' and word_lower in common_nouns:
        return 'NOUN'

    # If all lowercase and tagged as PROPN, likely wrong
    if pos == 'PROPN' and word == word_lower and not word[0].isupper():
        return 'NOUN'

    return pos


def analyze_english(word: str) -> list[dict]:
    """
    Analyze an English word and return morphological analyses.

    Uses spaCy for lemma and POS, MorphoLex for morphological segmentation,
    and MorphyNet for derivation type information.

    Args:
        word: English word to analyze

    Returns:
        List of dicts matching the lexicon.entries schema columns
    """
    results = []

    # Get spaCy analysis for lemma and POS
    lemma = word
    pos = ''

    if _nlp:
        doc = _nlp(word)
        if len(doc) > 0:
            token = doc[0]
            lemma = token.lemma_
            pos = token.pos_
            # Fix common POS tagging errors (Problem 5)
            pos = _fix_pos_tag(word, pos, lemma)

    # Load MorphoLex and MorphyNet data
    morpholex_data = _load_morpholex_data()
    morphynet_data = _load_morphynet_data()

    word_lower = word.lower()

    # Check if word is in MorphoLex
    morpholex_segm = morpholex_data.get(word_lower)

    if morpholex_segm:
        # Parse MorphoLex segmentation
        parsed = _parse_morpholex_segm(morpholex_segm)

        # Get derivation type from MorphyNet or infer from affixes
        morphynet_deriv = morphynet_data.get(word_lower)
        derivation_type = _get_derivation_type(
            parsed['prefixes'],
            parsed['suffixes'],
            morphynet_deriv
        )

        # Classify morphological type (Problem 2)
        morph_type = _classify_morph_type(
            parsed['root'],
            parsed['prefixes'],
            parsed['suffixes'],
            parsed['components']
        )

        # Build morphological features
        morphological_features = {}
        if parsed['prefixes']:
            morphological_features['prefixes'] = parsed['prefixes']
        if parsed['suffixes']:
            morphological_features['suffixes'] = parsed['suffixes']
        if parsed['components']:
            morphological_features['components'] = parsed['components']

        result = {
            'language_code': 'en',
            'word_native': word,
            'lemma': lemma,
            'root': parsed['root'],
            'pos': pos,
            'morph_type': morph_type,
            'derivation_type': derivation_type,
            'derived_from_root': parsed['root'],
            'derivation_mode': derivation_type,
            'compound_components': parsed['components'] if len(parsed['components']) > 1 else None,
            'morphological_features': morphological_features,
            'confidence': 1.0,
            'source_tool': 'morpholex+morphynet+spacy'
        }
        results.append(result)
    else:
        # Fallback: spaCy only - use lemma as root approximation
        result = {
            'language_code': 'en',
            'word_native': word,
            'lemma': lemma,
            'root': lemma,  # Use lemma as root when MorphoLex unavailable
            'pos': pos,
            'morph_type': 'UNKNOWN',
            'derivation_type': None,
            'derived_from_root': None,
            'derivation_mode': None,
            'compound_components': None,
            'morphological_features': {},
            'confidence': 0.5,
            'source_tool': 'spacy'
        }
        results.append(result)

    return results
